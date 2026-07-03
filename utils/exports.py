"""Export utilities for admin tables."""
import csv
import io
from fpdf import FPDF


def export_csv(filename, headers, rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    output.seek(0)
    return output.getvalue(), f'{filename}.csv', 'text/csv'


def export_excel(filename, headers, rows):
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Export'
        ws.append(headers)
        for row in rows:
            ws.append(list(row))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue(), f'{filename}.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    except ImportError:
        return export_csv(filename, headers, rows)


def export_pdf(filename, title, headers, rows):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, title, ln=True)
    pdf.ln(4)
    pdf.set_font('Helvetica', 'B', 9)
    col_width = max(25, (pdf.w - 20) / max(len(headers), 1))
    for h in headers:
        pdf.cell(col_width, 8, str(h)[:30], border=1)
    pdf.ln()
    pdf.set_font('Helvetica', '', 8)
    for row in rows:
        for cell in row:
            pdf.cell(col_width, 7, str(cell)[:40] if cell is not None else '', border=1)
        pdf.ln()
    data = pdf.output(dest='S')
    if isinstance(data, str):
        data = data.encode('latin-1')
    return data, f'{filename}.pdf', 'application/pdf'
